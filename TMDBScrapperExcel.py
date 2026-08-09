import requests
import openpyxl
import xlwings as xw
import time,os

global API_KEY

def get_api_key():
    # Check if the API key is already saved in a config file
    config_path = os.path.join(os.path.expanduser("~"), "Documents", "tmdbscraper_config.json")
    if os.path.exists(config_path):
        try:
            with open(config_path, "r") as f:
                data = json.load(f)
                api_key = data.get("api_key")
                if api_key:
                    return api_key
        except:
            pass

    # If not found, prompt the user to enter it
    print("Please enter your TMDB API key:")
    api_key = input("API Key: ").strip()

    # Save the API key for future use
    data = {"api_key": api_key}
    with open(config_path, "w") as f:
        json.dump(data, f, indent=4)

    return api_key

def get_movie_details_tmdb(title, year):
    
    # Step 1: Search for the movie
    search_url = "https://api.themoviedb.org/3/search/movie"
    params = {
        "api_key": API_KEY,
        "query": title,
        "year": year
    }
    search_response = requests.get(search_url, params=params).json()

    if not search_response["results"]:
        return None

    movie_id = search_response["results"][0]["id"]

    # Step 2: Get full movie details with cast, crew, images, etc.
    details_url = f"https://api.themoviedb.org/3/movie/{movie_id}"
    params = {
        "api_key": API_KEY,
        "append_to_response": "credits,images"
    }
    details = requests.get(details_url, params=params).json()

    # Extract director and writer
    director = "N/A"
    writer = "N/A"

    for crew_member in details["credits"]["crew"]:
        job = crew_member.get("job", "")
        name = crew_member.get("name", "")

        if job == "Director":
            director = name

        # Writers can appear under many job titles
        if job in ["Writer", "Screenplay", "Story", "Author", "Novel", "Characters"]:
            writer = name

    # Extract cast
    cast = []
    for actor in details["credits"]["cast"][:20]:  # limit to top 20
        cast.append({
            "actor_name": actor["name"],
            "character": actor["character"]
        })

    # Poster
    poster_url = (
        f"https://image.tmdb.org/t/p/w500{details['poster_path']}"
        if details.get("poster_path")
        else "N/A"
    )

    # Movie page URL
    movie_url = f"https://www.themoviedb.org/movie/{movie_id}"

    return {
        "Year": year,
        "Title": details.get("title", "N/A"),
        "Rating": details.get("vote_average", "N/A"),
        "Overview": details.get("overview", "N/A"),
        "Cast": cast,
        "Director": director,
        "Writer": writer,
        "URL": movie_url,
        "Image URL": poster_url
    }



# Function to read movie titles from Excel and update the details in the Excel file
def update_excel_with_movie_details(excel_file_path, sheet_name):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]
    count = 4
    DoEntry=False
    # Iterate over each row in column A
    for row in sheet.iter_rows(min_row=4, min_col=1, max_col=1):
        cell = row[0]
        movie_title = cell.value
        ColumnC=sheet["C"+str(count)].value
        if ColumnC == None:
            DoEntry=True
            time.sleep(3)  # This will pause the program for 3 seconds
            movie_details = get_movie_details_tmdb(movie_title, sheet["B"+str(count)].value if sheet["B"+str(count)].value else None)
        
            if movie_details:
                # Update the Excel sheet with the movie details
                sheet.cell(row=cell.row, column=2, value=movie_details['Year'])
                sheet.cell(row=cell.row, column=3, value=movie_details['Overview'])
                sheet.cell(row=cell.row, column=4, value=movie_details['Rating'])
                sheet.cell(row=cell.row, column=5, value=movie_details['Director'])
                sheet.cell(row=cell.row, column=6, value=movie_details['Writer'])
                sheet.cell(row=cell.row, column=7, value=movie_details['URL'])
                
                cast_text = "; ".join([f"{member['actor_name']} as {member['character']}" for member in movie_details['Cast']])
                sheet.cell(row=cell.row, column=8, value=cast_text)

                sheet.cell(row=cell.row, column=9, value=movie_details['Image URL'])

                if movie_details:
                    print(f"Title: {movie_details['Title']}")
                    print(f"Overview: {movie_details['Overview']}")
                    print(f"Rating: {movie_details['Rating']}")
                    print(f"Director: {movie_details['Director']}")
                    print(f"Writer: {movie_details['Writer']}")
                    print(f"URL: {movie_details['URL']}")
                    print(f"Image URL: {movie_details['Image URL']}")

                    print("\nCharacters:")
                    for member in movie_details['Cast']:
                        print(f" - {member['actor_name']} / {member['character']}")
                    workbook.save(excel_file_path)                        
        else:
            print("Skiped Movie "+str(count-3))                        
        count=count+1
        
    if DoEntry:
    # Save the updated Excel file
        
        # Use xlwings to add the VBA macro
        wb = xw.Book(excel_file_path)
        vba_code = r'''
            Sub AddButtons()
                Dim ws As Worksheet
                Set ws = ThisWorkbook.Sheets("Sheet1")
                Dim lastRow As Long
                lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row

                Dim btn As Button
                Dim i As Long
                For i = 4 To lastRow
                    Set btn = ws.Buttons.Add(ws.Cells(i, 10).Left, ws.Cells(i, 9).Top, 100, 20)
                    btn.Name = "btnShowImage" & i
                    btn.OnAction = "ShowImageAndInfo"
                    ' Set the button caption to the value in Column A
                    btn.Caption = ws.Cells(i, 1).Value
                Next i
            End Sub

        Sub ShowImage()
            ' In the Tools References dialog in Developer mode,
            ' scroll down and check the box for
            ' Microsoft Visual Basic for Applications Extensibility 5.3
            ' & Microsoft Forms 2.0 Object Library by inserting a form
            Dim btnName As String
            btnName = Application.Caller
            Dim ws As Worksheet
            Set ws = ThisWorkbook.Sheets("Sheet1")
            Dim rowIndex As Long
            rowIndex = ws.Shapes(btnName).TopLeftCell.Row

            Dim imageURL As String
            imageURL = ws.Cells(rowIndex, 9).Value

            If imageURL <> "N/A" Then
                ' Download the image from the URL
                Dim localFilePath As String
                localFilePath = DownloadImage(imageURL)
                
                If localFilePath <> "" Then
                    ' Create the UserForm dynamically
                    Dim VBComp As VBComponent
                    Set VBComp = ThisWorkbook.VBProject.VBComponents.Add(vbext_ct_MSForm)
                    With VBComp
                        .Properties("Width") = 320
                        .Properties("Height") = 390
                        .Properties("Caption") = ws.Cells(rowIndex, 2).Value
                    End With
                    
                    ' Add an Image control to the UserForm
                    Dim ImgControl As MSForms.Image
                    Set ImgControl = VBComp.Designer.Controls.Add("Forms.Image.1")
                    With ImgControl
                        .Left = 10
                        .Top = 10
                        .Width = 320
                        .Height = 350
                        .Picture = LoadPicture(localFilePath)
                    End With
                    
                    ' Show the dynamically created UserForm and delete it afterwards
                    With VBA.UserForms.Add(VBComp.Name)
                        .Show
                        ' Once the form is closed, remove it from the workbook
                        ThisWorkbook.VBProject.VBComponents.Remove VBComp
                    End With
                Else
                    MsgBox "Failed to download the image.", vbExclamation
                End If
            Else
                MsgBox "No image available for this movie.", vbExclamation
            End If
        End Sub
        
        Sub ShowImageAndInfo()
            Dim btnName As String
            btnName = Application.Caller
            
            Dim ws As Worksheet
            Set ws = ThisWorkbook.Sheets("Sheet1")
            
            Dim rowIndex As Long
            rowIndex = ws.Shapes(btnName).TopLeftCell.Row
            
            Dim imageURL As String
            imageURL = ws.Cells(rowIndex, 9).Value
            
            If imageURL <> "N/A" Then
                Dim localFilePath As String
                localFilePath = DownloadImage(imageURL)
                
                If localFilePath <> "" Then
                    
                    Dim VBComp As VBComponent
                    Set VBComp = ThisWorkbook.VBProject.VBComponents.Add(vbext_ct_MSForm)
                    With VBComp
                        .Properties("Width") = 400
                        .Properties("Height") = 600
                        .Properties("Caption") = ws.Cells(rowIndex, 2).Value
                    End With
                    
                    ' ---------------- IMAGE CONTROL ----------------
                    Dim ImgControl As MSForms.Image
                    Set ImgControl = VBComp.Designer.Controls.Add("Forms.Image.1")
                    
                    ImgControl.Picture = LoadPicture(localFilePath)
                    ImgControl.PictureSizeMode = fmPictureSizeModeZoom
                    ImgControl.Left = 10
                    ImgControl.Top = 10
                    ImgControl.Width = 380
                    ImgControl.Height = 300
                    
                    ' Store image path for full-screen viewer
                    ImgControl.Tag = localFilePath
                    
                    ' Add click handler for full-screen viewer
                    Dim CodeMod As CodeModule
                    Set CodeMod = VBComp.CodeModule
                    Dim LineNum As Long
                    LineNum = CodeMod.CountOfLines + 1
                    
                    CodeMod.InsertLines LineNum, _
                        "Private Sub " & ImgControl.Name & "_Click()" & VbCrLf & _
                        "    ShowFullScreenImage Me." & ImgControl.Name & ".Tag" & VbCrLf & _
                        "End Sub"
                    
                    ' ---------------- DESCRIPTION ----------------
                    Dim txtDesc As MSForms.TextBox
                    Set txtDesc = VBComp.Designer.Controls.Add("Forms.TextBox.1")
                    With txtDesc
                        .Text = "Description: " & ws.Cells(rowIndex, 3).Value
                        .Left = 10
                        .Top = 320
                        .Width = 380
                        .Height = 40
                        .Multiline = True
                        .WordWrap = True
                        .ScrollBars = fmScrollBarsVertical
                        .Locked = True
                        .BackColor = RGB(240, 240, 240)
                    End With
                    
                    ' ---------------- DIRECTOR ----------------
                    Dim lblDirector As MSForms.Label
                    Set lblDirector = VBComp.Designer.Controls.Add("Forms.Label.1")
                    With lblDirector
                        .Caption = "Director: " & ws.Cells(rowIndex, 5).Value
                        .Left = 10
                        .Top = 370
                        .Width = 380
                    End With
                    
                    ' ---------------- WRITER ----------------
                    Dim lblWriter As MSForms.Label
                    Set lblWriter = VBComp.Designer.Controls.Add("Forms.Label.1")
                    With lblWriter
                        .Caption = "Writer: " & ws.Cells(rowIndex, 6).Value
                        .Left = 10
                        .Top = 420
                        .Width = 380
                    End With
                    
                    ' ---------------- CAST ----------------
                    Dim txtCast As MSForms.TextBox
                    Set txtCast = VBComp.Designer.Controls.Add("Forms.TextBox.1")
                    With txtCast
                        .Text = "Cast: " & ws.Cells(rowIndex, 8).Value
                        .Left = 10
                        .Top = 460
                        .Width = 380
                        .Height = 70
                        .Multiline = True
                        .WordWrap = True
                        .ScrollBars = fmScrollBarsVertical
                        .Locked = True
                        .BackColor = RGB(240, 240, 240)
                    End With
                    
                    ' ---------------- LINK BUTTON ----------------
                    Dim btnLink As MSForms.CommandButton
                    Set btnLink = VBComp.Designer.Controls.Add("Forms.CommandButton.1")
                    With btnLink
                        .Caption = "Open Link"
                        .Left = 150
                        .Top = 540
                        .Width = 100
                        .Height = 30
                        .Tag = ws.Cells(rowIndex, 7).Value
                        .BackColor = RGB(0, 102, 204)
                        .ForeColor = RGB(255, 255, 255)
                        .Font.Bold = True
                        .Font.Size = 10
                    End With
                    
                    ' Add link button handler
                    LineNum = CodeMod.CountOfLines + 1
                    CodeMod.InsertLines LineNum, _
                        "Private Sub " & btnLink.Name & "_Click()" & VbCrLf & _
                        "    Dim link As String" & VbCrLf & _
                        "    link = Me." & btnLink.Name & ".Tag" & VbCrLf & _
                        "    If link <> """" Then" & VbCrLf & _
                        "        ThisWorkbook.FollowHyperlink link" & VbCrLf & _
                        "    Else" & VbCrLf & _
                        "        MsgBox ""No link available."", vbExclamation" & VbCrLf & _
                        "    End If" & VbCrLf & _
                        "End Sub"
                    
                    ' Show form
                    With VBA.UserForms.Add(VBComp.Name)
                        .Show
                        ThisWorkbook.VBProject.VBComponents.Remove VBComp
                    End With
                    
                Else
                    MsgBox "Failed to download the image.", vbExclamation
                End If
            Else
                MsgBox "No image available for this movie.", vbExclamation
            End If
        End Sub

        Public Sub ShowFullScreenImage(imagePath As String)
            If Dir(imagePath) = "" Then Exit Sub
            
            Dim VBComp As VBComponent
            Set VBComp = ThisWorkbook.VBProject.VBComponents.Add(vbext_ct_MSForm)
            
            With VBComp
                .Properties("Width") = Application.Width
                .Properties("Height") = Application.Height
                .Properties("Caption") = "Full Screen Image"
            End With
            
            Dim Img As MSForms.Image
            Set Img = VBComp.Designer.Controls.Add("Forms.Image.1")
            
            With Img
                .Picture = LoadPicture(imagePath)
                .PictureSizeMode = fmPictureSizeModeZoom
                .Left = 0
                .Top = 0
                .Width = Application.Width
                .Height = Application.Height
            End With
            
            With VBA.UserForms.Add(VBComp.Name)
                .Show
                ThisWorkbook.VBProject.VBComponents.Remove VBComp
            End With
        End Sub

        Function DownloadImage(ByVal url As String) As String
            Dim httpReq As Object
            Set httpReq = CreateObject("WinHttp.WinHttpRequest.5.1")
            httpReq.Open "GET", url, False
            httpReq.Send

            If httpReq.Status = 200 Then
                Dim stream As Object
                Set stream = CreateObject("ADODB.Stream")
                stream.Open
                stream.Type = 1 ' Binary
                stream.Write httpReq.responseBody
                stream.SaveToFile Environ("TEMP") & "\downloaded_image.jpg", 2 ' Overwrite if exists
                stream.Close
                DownloadImage = Environ("TEMP") & "\downloaded_image.jpg"
            Else
                DownloadImage = ""
            End If
        End Function
        
        Sub ShowInfoPopup()
             MsgBox "In the Tools References dialog in Developer mode, scroll down and check the box for Microsoft Visual Basic for Applications Extensibility 5.3 & Microsoft Forms 2.0 Object Library by inserting a form", vbInformation, "Info"
        End Sub
        
        Sub MakeSelfLinks_G_I()

            Dim ws As Worksheet
            Dim lastRow As Long
            Dim r As Long
            Dim v As String
            
            Set ws = ActiveSheet
            
            ' Find last row based on column G or I
            lastRow = ws.Cells(ws.Rows.Count, "G").End(xlUp).Row
            If ws.Cells(ws.Rows.Count, "I").End(xlUp).Row > lastRow Then
                lastRow = ws.Cells(ws.Rows.Count, "I").End(xlUp).Row
            End If
            
            For r = 4 To lastRow
                
                ' ----- Column G -----
                v = ws.Cells(r, "G").Value
                If v <> "" Then
                    ws.Hyperlinks.Add _
                        Anchor:=ws.Cells(r, "G"), _
                        Address:=v, _
                        TextToDisplay:=v
                End If
                
                ' ----- Column I -----
                v = ws.Cells(r, "I").Value
                If v <> "" Then
                    ws.Hyperlinks.Add _
                        Anchor:=ws.Cells(r, "I"), _
                        Address:=v, _
                        TextToDisplay:=v
                End If
                
            Next r

        End Sub
        '''

        wb.api.VBProject.VBComponents.Add(1).CodeModule.AddFromString(vba_code)
        
        ws = wb.sheets['Sheet1']
        btn = ws.api.Buttons().Add(0, 0, 50, 20)  # Left, Top, Width, Height
        btn.Name = "btnInfo"
        btn.OnAction = "ShowInfoPopup"
        btn.Caption = "Info"

        btn2= ws.api.Buttons().Add(75, 0, 100, 20)  # Left, Top, Width, Height
        btn2.Name = "btnAddButtons"
        btn2.OnAction = "AddButtons"
        btn2.Caption = "Add Buttons"

        # Save as .xlsm
        
        wb.save(excel_file_path.replace('.xlsx', '.xlsm'))
        wb.close()

import os
import json

config_path = os.path.join(os.path.expanduser("~"), "Documents", "tmdbscraper_config.json")

def load_saved_excel_path():
    if os.path.exists(config_path):
        try:
            with open(config_path, "r") as f:
                data = json.load(f)
                return data.get("excel_file_path")
        except:
            pass
    return None

def save_excel_path(path):
    # Load existing config if it exists
    try:
        with open(config_path, "r") as f:
            data = json.load(f)
    except FileNotFoundError:
        data = {"api_key": "TMDB_API_KEY"}  # Start fresh if no config exists

    # Update only the excel path
    data["excel_file_path"] = path

    # Save back without touching other keys
    with open(config_path, "w") as f:
        json.dump(data, f, indent=4)
    

excel_file_path = load_saved_excel_path()
API_KEY = get_api_key()

if excel_file_path is None:
    print("No saved Excel file path found.")
    print("Please enter the full path to your Excel file:")
    user_input = input("Excel file path: ").strip()

    # Expand ~ and environment variables
    user_input = os.path.expanduser(user_input)

    # If user entered a relative path, assume Documents
    if not os.path.isabs(user_input):
        user_input = os.path.join(os.path.expanduser("~"), "Documents", user_input)

    excel_file_path = user_input

    # Save it for next time
    save_excel_path(excel_file_path)
    print(f"Saved Excel path to {config_path}")

else:
    print(f"Loaded saved Excel path: {excel_file_path}")

sheet_name = "Sheet1"        
        
# # Example usage
# # excel_file_path = "My Movie Library.xlsx"
# excel_file_path = "F:/Richard/My Movie Library.xlsx"
# if not os.path.isabs(excel_file_path):
#    excel_file_path=os.path.join(os.path.expanduser("~"), "Documents")+"\\"+excel_file_path
# sheet_name = 'Sheet1'
update_excel_with_movie_details(excel_file_path, sheet_name)


